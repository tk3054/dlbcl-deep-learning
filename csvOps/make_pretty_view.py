#!/usr/bin/env python3
"""
Pretty-view exporter for a CSV.

Creates an .xlsx with a two-row header:
- Row 1: channel (dye removed)
- Row 2: stat (mean/median/etc.)
Means and medians are grouped and separated by borders.
"""

from pathlib import Path
from typing import Dict, List, Tuple

import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font


INTENSITY_STATS = {"mean", "median", "std", "min", "max", "intden", "rawintden"}
DYE_TOKENS = {
    "af647",
    "sparkviolet",
    "pacblue",
    "percp",
    "af594",
    "fitc",
}

# ============================================================================
# CONFIGURATION
# ============================================================================

INPUT_CSV = '/Users/taeeonkong/Desktop/DL Project/non-responder/01-03-2026 DLBCL 109241/sample1/combined_measurements.csv'
PRETTY_COLUMN_TOGGLES = {
    "_default": False,
    "global_cell_id": True,
    "unique_id": True,
    "sample": True,
    "image": True,
    "cell_id": True,
    "area": True,
    "x": False,
    "y": False,
    "circ": False,
    "ar": False,
    "round": False,
    "solidity": False,
}
PRETTY_KEEP_INTENSITY_STATS = ["mean", "median"]



def _split_channel_stat(col: str) -> Tuple[str, str]:
    parts = col.split("_")
    if len(parts) < 2:
        return col, ""
    stat = parts[-1]
    channel = "_".join(parts[:-1])
    return channel, stat


def _strip_dye_tokens(channel: str) -> str:
    parts = [p for p in channel.split("_") if p.lower() not in DYE_TOKENS]
    return "_".join(parts) if parts else channel


def _order_columns(columns: List[str]) -> Tuple[List[str], List[str], List[str], List[str]]:
    id_cols = []
    mean_cols = []
    median_cols = []
    other_intensity = []

    for col in columns:
        channel, stat = _split_channel_stat(col)
        if stat in INTENSITY_STATS and channel != col:
            if stat == "mean":
                mean_cols.append(col)
            elif stat == "median":
                median_cols.append(col)
            else:
                other_intensity.append(col)
        else:
            id_cols.append(col)

    mean_cols = sorted(mean_cols)
    median_cols = sorted(median_cols)
    other_intensity = sorted(other_intensity)
    return id_cols, mean_cols, median_cols, other_intensity


def export_pretty_view(input_csv: str, output_xlsx: str) -> Dict[str, str]:
    csv_path = Path(input_csv)
    if not csv_path.exists():
        return {"success": False, "error": f"CSV not found: {csv_path}"}

    df = pd.read_csv(csv_path)
    default_keep = PRETTY_COLUMN_TOGGLES.get("_default", True)
    keep_stats = set(PRETTY_KEEP_INTENSITY_STATS or [])
    filtered_cols = []

    for col in df.columns:
        channel, stat = _split_channel_stat(col)
        is_intensity = stat in INTENSITY_STATS and channel != col
        if is_intensity:
            if not keep_stats or stat in keep_stats:
                filtered_cols.append(col)
            continue
        if PRETTY_COLUMN_TOGGLES.get(col, default_keep):
            filtered_cols.append(col)

    df = df[filtered_cols]
    id_cols, mean_cols, median_cols, other_cols = _order_columns(list(df.columns))
    ordered_cols = id_cols + mean_cols + median_cols + other_cols
    df = df[ordered_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "pretty_view"

    header_row: List[str] = []
    expanded_rows: List[List[object]] = []
    id_cols_set = set(id_cols)
    intensity_cols_set = set(mean_cols + median_cols + other_cols)
    expanded_cols_set = id_cols_set | intensity_cols_set

    for col in ordered_cols:
        channel, stat = _split_channel_stat(col)
        if stat in INTENSITY_STATS and channel != col:
            channel_label = _strip_dye_tokens(channel)
            header_row.append(f"{channel_label} {stat}")
        else:
            header_row.append(col)

        if col in expanded_cols_set:
            header_row.append("")  # spacer for merged ID column

    ws.append(header_row)
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align

    for row in df.itertuples(index=False):
        expanded = []
        for col, value in zip(ordered_cols, row):
            channel, stat = _split_channel_stat(col)
            is_intensity = stat in INTENSITY_STATS and channel != col
            if is_intensity and value is not None:
                try:
                    value = int(round(float(value)))
                except (TypeError, ValueError):
                    pass
            expanded.append(value)
            if col in expanded_cols_set:
                expanded.append("")
        expanded_rows.append(expanded)

    for row in expanded_rows:
        ws.append(row)

    # Center align all cells (including header)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align

    def apply_group_borders(start_col: int, end_col: int) -> None:
        medium = Side(style="medium")
        for r in range(1, ws.max_row + 1):
            ws.cell(row=r, column=start_col).border = Border(left=medium)
            ws.cell(row=r, column=end_col).border = Border(right=medium)

    def count_expanded(cols: List[str]) -> int:
        count = 0
        for col in cols:
            count += 2 if col in expanded_cols_set else 1
        return count

    id_len = count_expanded(id_cols)
    mean_len = count_expanded(mean_cols)
    median_len = count_expanded(median_cols)

    if mean_len:
        apply_group_borders(id_len + 1, id_len + mean_len)
    if median_len:
        apply_group_borders(id_len + mean_len + 1, id_len + mean_len + median_len)

    # Merge header cells across spacer columns for expanded fields
    col_idx = 1
    for col in ordered_cols:
        if col in expanded_cols_set:
            ws.merge_cells(
                start_row=1,
                start_column=col_idx,
                end_row=1,
                end_column=col_idx + 1,
            )
            col_idx += 2
        else:
            col_idx += 1

    output_path = Path(output_xlsx)
    wb.save(output_path)
    return {"success": True, "output_xlsx": str(output_path)}


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a pretty-view XLSX from a CSV.")
    parser.add_argument(
        "input_csv",
        nargs="?",
        default=INPUT_CSV,
        help="Path to input CSV (or set INPUT_CSV in this file)",
    )
    parser.add_argument(
        "output_xlsx",
        nargs="?",
        default="",
        help="Path to output XLSX (optional; defaults next to input CSV)",
    )
    args = parser.parse_args()

    if not args.input_csv:
        print("✗ Please provide input_csv (arg or INPUT_CSV in this file).")
        return

    if not args.output_xlsx:
        input_path = Path(args.input_csv)
        args.output_xlsx = str(input_path.with_name(f"{input_path.stem}_pretty.xlsx"))

    result = export_pretty_view(args.input_csv, args.output_xlsx)
    if result.get("success"):
        print(f"✓ Wrote pretty view: {result['output_xlsx']}")
    else:
        print(f"✗ Pretty view export failed: {result.get('error')}")


if __name__ == "__main__":
    main()
